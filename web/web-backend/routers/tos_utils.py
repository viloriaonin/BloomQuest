import math
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BLOOM_LEVELS = [
    "Remember",
    "Understand",
    "Apply",
    "Analyze",
    "Evaluate",
    "Create"
]


DEFAULT_BLOOM_PERCENTAGES = {
    "Remember": 0.15,
    "Understand": 0.20,
    "Apply": 0.20,
    "Analyze": 0.20,
    "Evaluate": 0.15,
    "Create": 0.10,
}


SUPPORTED_QUESTION_TYPES = [
    "MCQ",
    "True or False",
    "Identification",
    "Essay",
    "Enumeration",
    "Matching Type",
    "Situational"
]


def normalize_question_types(question_types):
    """
    Removes duplicates while preserving order.
    """

    normalized = []

    for q in question_types:
        q = q.strip()

        if q not in normalized:
            normalized.append(q)

    if not normalized:
        normalized = ["MCQ"]

    return normalized


def compute_topic_weights(selected_topics, hours_dict):
    """
    Computes the percentage weight of every topic
    according to classroom hours.
    """

    total_hours = 0

    topic_hours = []

    for idx in selected_topics:

        hrs = float(hours_dict.get(str(idx), 1))

        topic_hours.append(hrs)

        total_hours += hrs

    if total_hours == 0:
        total_hours = len(topic_hours)

    weights = []

    for hrs in topic_hours:
        weights.append(hrs / total_hours)

    return weights


def distribute_items(total_items, weights):
    """
    Distributes question counts proportionally.
    """

    raw = [w * total_items for w in weights]

    rounded = [math.floor(x) for x in raw]

    remaining = total_items - sum(rounded)

    decimals = []

    for i, value in enumerate(raw):
        decimals.append((value - rounded[i], i))

    decimals.sort(reverse=True)

    for _, idx in decimals[:remaining]:
        rounded[idx] += 1

    return rounded


def compute_bloom_distribution(item_count):
    """
    Returns

    {
        Remember:2,
        Understand:2,
        ...
    }
    """

    raw = {}

    for bloom in BLOOM_LEVELS:

        raw[bloom] = DEFAULT_BLOOM_PERCENTAGES[bloom] * item_count

    bloom_counts = {}

    for b in BLOOM_LEVELS:

        bloom_counts[b] = math.floor(raw[b])

    remaining = item_count - sum(bloom_counts.values())

    decimals = []

    for b in BLOOM_LEVELS:
        decimals.append(
            (
                raw[b] - bloom_counts[b],
                b
            )
        )

    decimals.sort(reverse=True)

    for _, bloom in decimals[:remaining]:
        bloom_counts[bloom] += 1

    return bloom_counts


def allocate_question_types(bloom_counts, selected_types, start_pointer=0):
    """
    Assigns a question type to every generated question.

    IMPORTANT: `start_pointer` lets the caller carry the round-robin
    position across MULTIPLE calls (i.e. across topics). Previously this
    function always restarted its internal pointer at 0, which meant every
    topic's type-assignment began back at the front of `selected_types`.
    Since each topic usually only has a handful of items (Bloom counts
    split ~6-10 items six ways), the pointer rarely advanced far enough to
    reach types near the end of the list -- so on EVERY topic, the same
    later-listed types (e.g. Matching Type, Situational, Identification,
    depending on click order) were silently skipped every single time.
    Threading a running pointer through `compute_tos` fixes that: the
    round-robin now advances across the whole exam instead of restarting
    per topic, so every selected type gets a fair chance to appear.

    Returns (allocation_dict, next_pointer) so the caller can pass
    next_pointer into the next topic's call.
    """

    selected_types = normalize_question_types(selected_types)

    allocation = defaultdict(list)

    pointer = start_pointer

    for bloom in BLOOM_LEVELS:

        count = bloom_counts[bloom]

        for _ in range(count):

            allocation[bloom].append(

                selected_types[
                    pointer % len(selected_types)
                ]

            )

            pointer += 1

    return dict(allocation), pointer


def compute_tos(
    topics,
    selected_topic_indices,
    hours_dict,
    total_items,
    question_types,
):
    """
    Main TOS computation used by both

    • Excel generation

    • AI generation
    """

    selected_topics = [
        topics[i]
        for i in selected_topic_indices
    ]

    weights = compute_topic_weights(
        selected_topic_indices,
        hours_dict
    )

    items_per_topic = distribute_items(
        total_items,
        weights
    )

    results = []

    # Carried across every topic in this loop (see allocate_question_types
    # docstring above) instead of resetting to 0 per topic.
    type_pointer = 0

    for topic, hrs, items in zip(
        selected_topics,
        [
            float(hours_dict.get(str(i), 1))
            for i in selected_topic_indices
        ],
        items_per_topic
    ):

        bloom_counts = compute_bloom_distribution(items)

        q_distribution, type_pointer = allocate_question_types(
            bloom_counts,
            question_types,
            start_pointer=type_pointer,
        )

        results.append({

            "topic_name": topic["name"],

            "ilo": topic.get("ilo", ""),

            "ilo_description": topic.get("ilo_description", ""),

            "ilo_num": "",

            "hours_a": hrs,

            "minutes_b": 2,

            "weight": round(
                hrs /
                sum(
                    float(hours_dict.get(str(i), 1))
                    for i in selected_topic_indices
                ) * 100,
                2
            ),

            "items": items,

            "bloom_counts": bloom_counts,

            "question_distribution": q_distribution

        })
        
    return results

def generate_tos_from_institutional_template(selected_topics_data, course_code, course_title, whole_total_points):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TOS"
    ws.views.sheetView[0].showGridLines = True

    # Institutional Brand Visual Formats
    font_header_title = Font(name="Calibri", size=11, bold=True)
    font_main_label = Font(name="Calibri", size=11, bold=True)
    font_body_data = Font(name="Calibri", size=11)
    
    thin_border_side = Side(style='thin', color='000000')
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Establish Institutional Identity Header Section
    ws['B7'] = "Republic of the Philippines"
    ws['B8'] = "BATANGAS STATE UNIVERSITY"
    ws['B9'] = "The National Engineering University"
    ws['B10'] = "Lipa Campus"
    ws['B11'] = "A. Tanco Drive, Marawoy, Lipa City, Batangas , Philippines 4217"
    ws['B12'] = "Tel Nos. : (+63 43) 980-0385; 980-0392 to local 3130"
    ws['B13'] = "E-mail Address: cics.lipa@g.batstate-u.edu.ph | Website Address: http://www.batstate-u.edu.ph"
    ws['B14'] = "                         College of Informatics and Computing Sciences"
    ws['B15'] = "TABLE OF SPECIFICATIONS\nFinal Examination\nFirst Semester, AY 2026 – 2027"
    ws['B15'].alignment = Alignment(wrap_text=True)

    ws['B18'] = f"COURSE CODE : {course_code or 'IT 332'}"
    ws['B19'] = f"COURSE TITLE: {course_title or 'Integrative Programming and Technologies'}"
    
    # Layout the Double-Row Split Column Headers (Rows 20 to 22)
    ws.merge_cells("G20:R20")
    ws['G20'] = " Indicate the test items that correspond to the following levels of Intended Learning Outcomes"
    ws['G20'].alignment = Alignment(horizontal="center")
    
    ws['B21'] = "TOPICS"
    ws['C21'] = "*ILOs"
    ws['D21'] = "NO. OF HRS "
    ws['F21'] = "WEIGHT (%) **"
    ws['S21'] = "TOTAL NO. OF POINTS"
    
    bloom_headers = ["REMEMBER", "UNDERSTAND", "APPLY ", "ANALYZE", "EVALUATE", "CREATE"]
    for idx, b_lvl in enumerate(bloom_headers):
        col_start = 7 + (idx * 2)
        ws.cell(row=21, column=col_start, value=b_lvl)
        ws.cell(row=21, column=col_start + 1, value="%")
        
    ws['D22'] = "A*"
    ws['E22'] = "B*"

    # Internal keys must match what tos_service.py actually produces
    # (BLOOMS_LEVELS = Remember/Understand/Apply/Analyze/Evaluate/Create).
    # Display labels are separate -- the template wants uppercase headers,
    # but the lookup key into topic["bloom_counts"] must match the caller.
    BLOOMS_LEVELS = ["Remember", "Understand", "Apply", "Analyze", "Evaluate", "Create"]
    
    for r in [20, 21, 22]:
        for c in range(2, 20):
            cell = ws.cell(row=r, column=c)
            cell.font = font_header_title
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = grid_border

    start_row = 23
    num_topics = len(selected_topics_data)
    total_row_index = start_row + num_topics
    
    for i, topic in enumerate(selected_topics_data):
        current_row = start_row + i
        
        ws.cell(row=current_row, column=2, value=topic["topic_name"]).alignment = Alignment(horizontal="left")
        ws.cell(row=current_row, column=3, value=topic.get("ilo", f"ILO {topic.get('ilo_num', 1)}"))
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(row=current_row, column=4, value=float(topic["hours_a"]))

        # Column B* = "minutes allotted to answer the test item/s" per the
        # template's own footnote -- this is a literal per-topic value, NOT
        # a formula derived from hours share (that was the original bug).
        ws.cell(row=current_row, column=5, value=float(topic.get("minutes_b", 2.0)))
        ws.cell(row=current_row, column=6, value=f"=IFERROR(S{current_row}/$S${total_row_index}*100,0)")

        for idx in range(6):
            item_col = 7 + (idx * 2)
            pct_col = item_col + 1
            ws.cell(row=current_row, column=item_col, value=topic.get("bloom_counts", {}).get(BLOOMS_LEVELS[idx], 0))
            ws.cell(row=current_row, column=pct_col, value=f"=IFERROR(({get_column_letter(item_col)}{current_row}/$S${total_row_index})*100,0)")

        ws.cell(row=current_row, column=19, value=f"=SUM(G{current_row},I{current_row},K{current_row},M{current_row},O{current_row},Q{current_row})")
        
        for c in range(2, 20):
            cell = ws.cell(row=current_row, column=c)
            cell.font = font_body_data
            cell.border = grid_border
            if c >= 4:
                cell.alignment = Alignment(horizontal="right")

    # Bottom Aggregate row configurations
    ws.cell(row=total_row_index, column=2, value="Total").font = font_main_label
    ws.cell(row=total_row_index, column=4, value=f"=SUM(D23:D{total_row_index-1})")
    ws.cell(row=total_row_index, column=5, value=f"=SUM(E23:E{total_row_index-1})")
    ws.cell(row=total_row_index, column=6, value=f"=SUM(F23:F{total_row_index-1})")
    
    for idx in range(6):
        item_col = 7 + (idx * 2)
        pct_col = item_col + 1
        ws.cell(row=total_row_index, column=item_col, value=f"=SUM({get_column_letter(item_col)}23:{get_column_letter(item_col)}{total_row_index-1})")
        ws.cell(row=total_row_index, column=pct_col, value=f"=SUM({get_column_letter(pct_col)}23:{get_column_letter(pct_col)}{total_row_index-1})")
        
    ws.cell(row=total_row_index, column=19, value=f"=SUM(S23:S{total_row_index-1})")
    
    for c in range(2, 20):
        cell = ws.cell(row=total_row_index, column=c)
        cell.font = font_main_label
        cell.border = grid_border
        if c >= 4:
            cell.alignment = Alignment(horizontal="right")

    # Bottom Signature Deck Content
    sign_row = total_row_index + 2
    ws.cell(row=sign_row, column=2, value="Prepared by:")
    ws.cell(row=sign_row, column=9, value="Checked and Verified by:")
    ws.cell(row=sign_row, column=15, value="Approved by:")
    
    name_row = sign_row + 2
    ws.cell(row=name_row, column=2, value="Faculty Instructor").font = font_main_label
    ws.cell(row=name_row, column=9, value="Mr. DIONECES O. ALIMOREN").font = font_main_label
    ws.cell(row=name_row, column=15, value="Dr. RYNDEL V. AMORADO").font = font_main_label

    # Legend footnotes, matching the original template
    legend_row = name_row + 3
    ws.cell(row=legend_row, column=2, value="*ILO - Intended Learning Outcomes")
    ws.cell(row=legend_row + 1, column=2, value="*A - No. of hours the topic was covered in class")
    ws.cell(row=legend_row + 2, column=2, value="*B - No. of minutes alloted to answer the test item/s")
    ws.cell(row=legend_row + 3, column=2, value="**Weight (%) = (no. of  points for a given topic /total no. of points)* 100")

    # whole_total_points is the target the caller confirmed in the Step 3 TOS
    # preview. The sheet's own formulas compute the actual total independently
    # (S{total_row_index}) -- if they disagree, something upstream (e.g. a
    # manual edit that wasn't re-validated) let a mismatched matrix through.
    # Flag it loudly rather than shipping a TOS that quietly doesn't add up.
    actual_total = sum(
        sum(t.get("bloom_counts", {}).get(level, 0) for level in BLOOMS_LEVELS)
        for t in selected_topics_data
    )
    if whole_total_points and actual_total != whole_total_points:
        warning_cell = ws.cell(
            row=legend_row + 5, column=2,
            value=(f"⚠ WARNING: Bloom's item counts sum to {actual_total}, "
                   f"but target Total No. of Points was {whole_total_points}. "
                   f"Re-check the TOS matrix before distributing this file."),
        )
        warning_cell.font = Font(name="Calibri", size=11, bold=True, color="CC0000")

    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['S'].width = 24
    
    return wb